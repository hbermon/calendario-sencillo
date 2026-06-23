import os
import sys
import tkinter as tk
from tkinter import ttk
from datetime import datetime
from dateutil.relativedelta import relativedelta
from tkcalendar import Calendar

def resolver_ruta(ruta_relativa):
    """ Obtiene la ruta absoluta para recursos empaquetados por PyInstaller """
    try:
        # PyInstaller crea una carpeta temporal y guarda la ruta en _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, ruta_relativa)

class CalendarioEventosApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Calendario de Eventos")
        
        # Configuración de ventana sin bordes (Estilo Widget Windows 11)
        self.overrideredirect(True)
        self.attributes("-topmost", True)
        self.configure(bg="#f3f3f3")

        # Banderas de control
        self.bloqueo_sincronizacion = False
        self.esta_fijado = False  # Estado inicial: No fijado (Cierre inteligente activo)

        # 1. Cargar Configuración de Colores y Eventos desde el mismo Excel
        self.colores_por_tipo = self.cargar_configuracion_colores()
        self.eventos = self.cargar_eventos()

        # Contenedor principal
        self.main_frame = ttk.Frame(self, padding=10)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # --- BOTÓN DE PIN / CHINCHETA (Esquina Superior Derecha) ---
        top_bar = tk.Frame(self.main_frame, bg="#f3f3f3")
        top_bar.grid(row=0, column=0, columnspan=2, sticky="ne", padx=5)
        
        self.btn_pin = tk.Button(
            top_bar, text="📍 Fijar", font=("Segoe UI", 8, "bold"),
            bg="#e0e0e0", fg="#333333", bd=0, padx=8, pady=2,
            cursor="hand2", command=self.conmutar_fijacion
        )
        self.btn_pin.pack(anchor="ne")

        hoy = datetime.today()
        siguiente_mes = hoy + relativedelta(months=1)

        # --- CALENDARIO 1: MES ACTUAL (IZQUIERDA) ---
        self.cal_actual = Calendar(
            self.main_frame, selectmode='day', 
            locale='es_ES', firstweekday='sunday',
            year=hoy.year, month=hoy.month, day=hoy.day,
            background='#2c3e50', foreground='white', 
            headersbackground='#34495e', headersforeground='white',
            normalbackground='white', normalforeground='black',
            weekendbackground='#f5f5f5', weekendforeground='black'
        )
        self.cal_actual.grid(row=1, column=0, padx=10, pady=10)

        # --- CALENDARIO 2: MES SIGUIENTE (DERECHA) ---
        self.cal_siguiente = Calendar(
            self.main_frame, selectmode='day', 
            locale='es_ES', firstweekday='sunday',
            year=siguiente_mes.year, month=siguiente_mes.month, day=1,
            background='#2c3e50', foreground='white', 
            headersbackground='#34495e', headersforeground='white',
            normalbackground='white', normalforeground='black',
            weekendbackground='#f5f5f5', weekendforeground='black'
        )
        self.cal_siguiente.grid(row=1, column=1, padx=10, pady=10)

        # 2. LEYENDA DE COLORES (Parte Inferior)
        self.crear_leyenda_inferior()

        # Inyección de sincronización de navegación bidireccional mediante API Oficial
        self.inyectar_navegacion_sincronizada()

        # Dibujar marcas y eventos iniciales
        self.refrescar_toda_la_interfaz()

        # Crear el Tooltip flotante para el Hover
        self.tooltip = tk.Toplevel(self)
        self.tooltip.withdraw()
        self.tooltip.overrideredirect(True)
        self.tooltip_label = tk.Label(self.tooltip, bg="#2d2d2d", fg="white", 
                                      font=("Segoe UI", 9), padx=5, pady=3, 
                                      relief=tk.SOLID, borderwidth=1)
        self.tooltip_label.pack()

        # Vincular eventos de Mouse (Hover y Clic)
        self.bind_eventos_mouse(self.cal_actual)
        self.bind_eventos_mouse(self.cal_siguiente)
        
        # 3. Habilitar arrastre de la ventana usando el fondo del main_frame
        self.main_frame.bind("<Button-1>", self.iniciar_arrastre)
        self.main_frame.bind("<B1-Motion>", self.ejecutar_arrastre)

        # Ubicar la ventana cerca de la esquina superior izquierda al arrancar
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        self.geometry(f'{width}x{height}+50+50')

        # Cierre inteligente al perder el foco (Solo si no está fijado)
        self.bind("<FocusOut>", self.evaluar_cierre)
        self.after(100, self.focus_force)

    def conmutar_fijacion(self):
        """Alterna el estado de fijación del widget en primer plano"""
        self.esta_fijado = not self.esta_fijado
        if self.esta_fijado:
            self.btn_pin.config(text="📌 Fijado", bg="#3498db", fg="white")
            self.attributes("-topmost", True)
        else:
            self.btn_pin.config(text="📍 Fijar", bg="#e0e0e0", fg="#333333")
            self.focus_force()

    def cargar_configuracion_colores(self):
        excel_path = 'eventos.xlsx'
        colores_defecto = {
            'Fin sprint': '#3498db',
            'Inicio sprint': '#2ecc71',
            'Festivo': '#e74c3c',
            'Predeterminado': '#95a5a6'
        }
        
        from openpyxl import load_workbook, Workbook
        if not os.path.exists(excel_path):
            wb = Workbook()
            ws_ev = wb.active
            ws_ev.title = "Eventos"
            ws_ev.append(['fecha', 'evento', 'tipo'])
            ws_ev.append([datetime.today().strftime('%Y-%m-%d'), 'Evento de Prueba', 'Inicio sprint'])
            ws_cfg = wb.create_sheet(title="Configuracion")
            ws_cfg.append(['tipo_evento', 'codigo_hex'])
            for k, v in colores_defecto.items():
                ws_cfg.append([k, v])
            wb.save(excel_path)
            return colores_defecto

        try:
            wb = load_workbook(excel_path)
            if "Configuracion" not in wb.sheetnames:
                ws_cfg = wb.create_sheet(title="Configuracion")
                ws_cfg.append(['tipo_evento', 'codigo_hex'])
                for k, v in colores_defecto.items():
                    ws_cfg.append([k, v])
                wb.save(excel_path)
                return colores_defecto
            
            ws = wb["Configuracion"]
            colores_dict = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    colores_dict[str(row[0]).strip()] = str(row[1]).strip()
            
            if 'Predeterminado' not in colores_dict:
                colores_dict['Predeterminado'] = '#95a5a6'
                
            wb.close()
            return colores_dict
        except Exception as e:
            print(f"Error al cargar configuración de colores: {e}")
            return colores_defecto

    def crear_leyenda_inferior(self):
        leyenda_frame = tk.Frame(self.main_frame, bg="#f3f3f3", pady=5)
        leyenda_frame.grid(row=2, column=0, columnspan=2, sticky="ew")
        
        leyenda_frame.columnconfigure(0, weight=1)
        container = tk.Frame(leyenda_frame, bg="#f3f3f3")
        container.pack(anchor="center")

        lbl_color_hoy = tk.Label(container, bg="#f1c40f", width=2, height=1, relief=tk.SOLID, bd=1)
        lbl_color_hoy.pack(side=tk.LEFT, padx=(5, 2))
        lbl_texto_hoy = tk.Label(container, text="Hoy", bg="#f3f3f3", font=("Segoe UI", 8, "bold"), fg="black")
        lbl_texto_hoy.pack(side=tk.LEFT, padx=(0, 15))

        for tipo, hex_color in self.colores_por_tipo.items():
            if tipo == 'Predeterminado': continue
            
            lbl_color = tk.Label(container, bg=hex_color, width=2, height=1, relief=tk.SOLID, bd=1)
            lbl_color.pack(side=tk.LEFT, padx=(5, 2))
            
            lbl_texto = tk.Label(container, text=tipo, bg="#f3f3f3", font=("Segoe UI", 8), fg="black")
            lbl_texto.pack(side=tk.LEFT, padx=(0, 15))

    def inyectar_navegacion_sincronizada(self):
        """ Sincronización reactiva basada en eventos utilizando la API correcta (see) """
        
        def sincronizar_desde_izquierdo(event):
            if self.bloqueo_sincronizacion: return
            self.bloqueo_sincronizacion = True
            
            # Leer qué mes quedó en la izquierda
            mes, anio = self.cal_actual.get_displayed_month()
            # Calcular el mes siguiente para la derecha
            fecha_der = datetime(anio, mes, 1) + relativedelta(months=1)
            # Método correcto de tkcalendar para posicionar la vista
            self.cal_siguiente.see(fecha_der)
            
            self.refrescar_toda_la_interfaz()
            self.bloqueo_sincronizacion = False

        def sincronizar_desde_derecho(event):
            if self.bloqueo_sincronizacion: return
            self.bloqueo_sincronizacion = True
            
            # Leer qué mes quedó en la derecha
            mes, anio = self.cal_siguiente.get_displayed_month()
            # Calcular el mes anterior para la izquierda
            fecha_izq = datetime(anio, mes, 1) - relativedelta(months=1)
            # Método correcto de tkcalendar para posicionar la vista
            self.cal_actual.see(fecha_izq)
            
            self.refrescar_toda_la_interfaz()
            self.bloqueo_sincronizacion = False

        # Escuchar de forma reactiva el evento oficial de cambio de mes
        self.cal_actual.bind("<<CalendarMonthChanged>>", sincronizar_desde_izquierdo)
        self.cal_siguiente.bind("<<CalendarMonthChanged>>", sincronizar_desde_derecho)

    def cargar_eventos(self):
        """Lee la pestaña 'Eventos' con la nueva estructura de rangos (fecha_inicio, fecha_fin)"""
        excel_path = 'eventos.xlsx'
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            nombre_hoja = "Eventos" if "Eventos" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[nombre_hoja]
            
            eventos_dict = {}
            
            # Recorrer filas. Estructura esperada: col 0=inicio, col 1=fin, col 2=evento, col 3=tipo
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: 
                    continue
                
                # Parseo seguro de Fecha Inicio
                if isinstance(row[0], datetime):
                    dt_inicio = row[0]
                else:
                    dt_inicio = datetime.strptime(str(row[0]).split(" ")[0], '%Y-%m-%d')
                
                # Parseo seguro de Fecha Fin (Opcional)
                dt_fin = None
                if row[1]:
                    if isinstance(row[1], datetime):
                        dt_fin = row[1]
                    else:
                        try:
                            dt_fin = datetime.strptime(str(row[1]).split(" ")[0], '%Y-%m-%d')
                        except:
                            dt_fin = None

                texto_evento = str(row[2]) if row[2] else ""
                tipo_evento = str(row[3]).strip() if row[3] else "Predeterminado"
                
                # Determinar los días a marcar
                lista_dias = []
                if dt_fin and dt_fin > dt_inicio:
                    # Es un rango (Sprint u otro bloque)
                    aux_dt = dt_inicio
                    while aux_dt <= dt_fin:
                        lista_dias.append(aux_dt.strftime('%Y-%m-%d'))
                        aux_dt += relativedelta(days=1)
                else:
                    # Es un evento de un solo día (Festivo, Hito)
                    lista_dias.append(dt_inicio.strftime('%Y-%m-%d'))

                # Guardar en el diccionario estructurado
                for celda_fecha in lista_dias:
                    if celda_fecha not in eventos_dict:
                        eventos_dict[celda_fecha] = []
                    
                    eventos_dict[celda_fecha].append({
                        'texto': texto_evento,
                        'tipo': tipo_evento,
                        'es_rango': dt_fin is not None
                    })
                    
            wb.close()
            return eventos_dict
        except Exception as e:
            print(f"Error al cargar el archivo Excel estructurado: {e}")
            return {}

    def bind_eventos_mouse(self, calendario):
        calendario.bind("<Motion>", lambda event: self.verificar_hover(event, calendario))
        calendario.bind("<Leave>", lambda event: self.ocultar_tooltip())
        calendario.bind("<<CalendarSelected>>", lambda event: self.mostrar_evento_seleccionado(calendario))

    def verificar_hover(self, event, calendario):
        x, y = event.x, event.y
        region = calendario.identify(x, y)
        
        if region == "day":
            date = calendario.get_date_at(x, y)
            if date:
                fecha_str = date.strftime('%Y-%m-%d')
                hoy_str = datetime.today().strftime('%Y-%m-%d')
                
                textos = []
                if fecha_str == hoy_str:
                    textos.append("★ Hoy")
                if fecha_str in self.eventos:
                    textos.extend([f"• {ev['texto']} ({ev['tipo']})" for ev in self.eventos[fecha_str]])
                
                if textos:
                    tooltip_text = "\n".join(textos)
                    gx = self.winfo_pointerx() + 15
                    gy = self.winfo_pointery() + 15
                    self.tooltip_label.config(text=tooltip_text)
                    self.tooltip.geometry(f"+{gx}+{gy}")
                    self.tooltip.deiconify()
                    return
        self.ocultar_tooltip()

    def ocultar_tooltip(self):
        self.tooltip.withdraw()

    def mostrar_evento_seleccionado(self, calendario):
        """Muestra el popup de detalles integrando el icono institucional de la app"""
        try:
            fecha_dt = calendario.selection_get()
        except:
            fecha_dt = None

        if not fecha_dt: 
            return

        # Validar si el día seleccionado pertenece al mes real de ese calendario
        mes_visible, ano_visible = calendario.get_displayed_month()
        if fecha_dt.month != mes_visible or fecha_dt.year != ano_visible:
            self.cal_actual.selection_set(None)
            self.cal_siguiente.selection_set(None)
            return

        fecha_str = fecha_dt.strftime('%Y-%m-%d')
        if fecha_str not in self.eventos: 
            self.cal_actual.selection_set(None)
            self.cal_siguiente.selection_set(None)
            return

        textos = [f"• {ev['texto']} ({ev['tipo']})" for ev in self.eventos[fecha_str]]
        detalle_eventos = "\n".join(textos)

        # Crear Ventana Emergente (Popup)
        popup = tk.Toplevel(self)
        popup.title("Detalle de Eventos")
        popup.configure(bg="#2c3e50")
        popup.geometry("340x220")
        popup.resizable(False, False)
        
        # --- CONFIGURACIÓN DEL ICONO DE LA VENTANA ---
        try:
            # Usamos la función de rescate para PyInstaller
            ruta_icono = resolver_ruta("calendario.ico")
            popup.iconbitmap(ruta_icono)
        except Exception as e:
            pass
        
        # Centrar el popup respecto a la app principal
        px = self.winfo_x() + (self.winfo_width() // 2) - 170
        py = self.winfo_y() + (self.winfo_height() // 2) - 110
        popup.geometry(f"+{px}+{py}")
        popup.attributes("-topmost", True)

        # Título de la fecha
        lbl_fecha = tk.Label(popup, text=f"Eventos: {fecha_dt.strftime('%d/%m/%Y')}", 
                             font=("Segoe UI", 11, "bold"), bg="#2c3e50", fg="#f1c40f", pady=8)
        lbl_fecha.pack(side=tk.TOP, fill=tk.X)

        # --- LÓGICA DE CIERRE SEGURO ---
        def ejecutar_cierre_limpio():
            try:
                self.cal_actual.selection_set(None)
                self.cal_siguiente.selection_set(None)
            except:
                pass
            self.refrescar_toda_la_interfaz()
            popup.destroy()

        # Botón de cierre
        btn_cerrar = tk.Button(popup, text="Cerrar", bg="#34495e", fg="white", 
                               font=("Segoe UI", 9, "bold"), activebackground="#1a252f", 
                               activeforeground="white", bd=0, cursor="hand2", 
                               command=ejecutar_cierre_limpio)
        btn_cerrar.pack(side=tk.BOTTOM, pady=12, ipadx=20, ipady=3)

        # Caja de Texto
        txt_box = tk.Text(popup, bg="#34495e", fg="white", font=("Segoe UI", 10), wrap=tk.WORD, bd=0, padx=10, pady=5)
        txt_box.insert(tk.END, detalle_eventos)
        txt_box.config(state=tk.DISABLED)
        txt_box.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=15, pady=2)

        popup.protocol("WM_DELETE_WINDOW", ejecutar_cierre_limpio)
        popup.focus_set()

    def refrescar_toda_la_interfaz(self):
        """
        Dibuja marcas aplicando prioridades y alternancia cronológica de sprints.
        Garantiza que si un FESTIVO cae sábado o domingo, se pinte correctamente, 
        mientras que los fines de semana ordinarios se mantienen limpios.
        """
        # 1. Limpiar por completo ambos calendarios
        self.cal_actual.calevent_remove('all')
        self.cal_siguiente.calevent_remove('all')
        
        hoy = datetime.today()
        hoy_str = hoy.strftime('%Y-%m-%d')
        
        # Marcar siempre el día de 'Hoy' (Prioridad Dorada)
        for cal in [self.cal_actual, self.cal_siguiente]:
            cal.calevent_create(hoy, 'Hoy', 'tag_hoy')
            cal.tag_config('tag_hoy', background='#f1c40f', foreground='black')

        # --- LÓGICA DE ALTERNANCIA CRONOLÓGICA DE SPRINTS ---
        dias_sprint_ordenados = sorted([
            f for f, eventos in self.eventos.items() 
            if any('sprint' in ev['tipo'].lower() for ev in eventos)
        ])

        mapeo_color_sprint = {}
        if dias_sprint_ordenados:
            indice_bloque_sprint = 0
            fecha_anterior = datetime.strptime(dias_sprint_ordenados[0], '%Y-%m-%d')
            mapeo_color_sprint[dias_sprint_ordenados[0]] = indice_bloque_sprint
            
            for i in range(1, len(dias_sprint_ordenados)):
                fecha_actual = datetime.strptime(dias_sprint_ordenados[i], '%Y-%m-%d')
                diferencia = (fecha_actual - fecha_anterior).days
                
                texto_ant = next(ev['texto'] for ev in self.eventos[fecha_anterior.strftime('%Y-%m-%d')] if 'sprint' in ev['tipo'].lower())
                texto_act = next(ev['texto'] for ev in self.eventos[fecha_actual.strftime('%Y-%m-%d')] if 'sprint' in ev['tipo'].lower())
                
                if diferencia > 3 or texto_ant != texto_act:
                    indice_bloque_sprint += 1
                
                mapeo_color_sprint[dias_sprint_ordenados[i]] = indice_bloque_sprint
                fecha_anterior = fecha_actual

        # Colores de Sprints
        color_sprint_a = self.colores_por_tipo.get('Inicio sprint', '#2ecc71')
        color_sprint_b = self.colores_por_tipo.get('Fin sprint', '#3498db')
        color_sprint_a_intenso = '#1b7e43' 
        color_sprint_b_intenso = '#1f5f8a' 

        def oscurecer_color(hex_color, factor=0.7):
            hex_color = hex_color.lstrip('#')
            if len(hex_color) != 6: return '#444444'
            try:
                r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                return f'#{int(r*factor):02x}{int(g*factor):02x}{int(b*factor):02x}'
            except: return '#444444'

        mes_izq, ano_izq = self.cal_actual.get_displayed_month()
        mes_der, ano_der = self.cal_siguiente.get_displayed_month()

        # --- RECORRER Y PINTAR CADA DÍA ---
        for fecha_str, lista_eventos in self.eventos.items():
            if fecha_str == hoy_str:
                continue  
                
            try:
                fecha_dt = datetime.strptime(fecha_str, '%Y-%m-%d')
                
                # 2. EVALUAR PRIORIDAD (Festivos > Sprints > Otros)
                evento_ganador = lista_eventos[0]
                es_festivo = False
                for ev in lista_eventos:
                    if 'festivo' in ev['tipo'].lower():
                        evento_ganador = ev
                        es_festivo = True
                        break
                
                # --- NUEVA REGLA DE FIN DE SEMANA BLINDADO ---
                # Si es sábado (5) o domingo (6) y NO es un festivo, se ignora el pintado ordinario.
                if fecha_dt.weekday() in [5, 6] and not es_festivo:
                    continue

                tipo = evento_ganador['tipo']
                texto = evento_ganador['texto']
                
                # 3. DETERMINAR SI CORRESPONDE APLICAR COINCIDENCIA (OSCURECIMIENTO)
                if es_festivo:
                    hay_coincidencia = False
                else:
                    hay_coincidencia = len(lista_eventos) > 1

                # 4. ASIGNACIÓN DEL COLOR DE FONDO FINAL
                if 'sprint' in tipo.lower():
                    bloque = mapeo_color_sprint.get(fecha_str, 0)
                    if hay_coincidencia:
                        color_fondo = color_sprint_a_intenso if bloque % 2 == 0 else color_sprint_b_intenso
                    else:
                        color_fondo = color_sprint_a if bloque % 2 == 0 else color_sprint_b
                else:
                    color_base = self.colores_por_tipo.get(tipo, self.colores_por_tipo.get('Predeterminado', '#95a5a6'))
                    if hay_coincidencia:
                        color_fondo = oscurecer_color(color_base, factor=0.7)
                    else:
                        color_fondo = color_base 

                # 5. INYECTAR EXCLUSIVAMENTE EN EL MES NATIVO
                tag_name = f"tag_{fecha_str}"
                
                if fecha_dt.month == mes_izq and fecha_dt.year == ano_izq:
                    self.cal_actual.calevent_create(fecha_dt, texto, tag_name)
                    self.cal_actual.tag_config(tag_name, background=color_fondo, foreground='white')
                    
                if fecha_dt.month == mes_der and fecha_dt.year == ano_der:
                    self.cal_siguiente.calevent_create(fecha_dt, texto, tag_name)
                    self.cal_siguiente.tag_config(tag_name, background=color_fondo, foreground='white')
                    
            except ValueError:
                continue

    def iniciar_arrastre(self, event):
        self.x_offset = event.x
        self.y_offset = event.y

    def ejecutar_arrastre(self, event):
        x = self.winfo_x() + (event.x - self.x_offset)
        y = self.winfo_y() + (event.y - self.y_offset)
        self.geometry(f"+{x}+{y}")

    def evaluar_cierre(self, event):
        if self.esta_fijado:
            return
        self.after(120, self._verificar_foco_real)

    def _verificar_foco_real(self):
        focus_widget = self.focus_get()
        if focus_widget is None:
            ventanas_secundarias = [w for w in self.winfo_children() if isinstance(w, tk.Toplevel) and w.winfo_viewable()]
            ventanas_activas = [v for v in ventanas_secundarias if v != self.tooltip]
            if not ventanas_activas:
                self.destroy()

if __name__ == "__main__":
    app = CalendarioEventosApp()
    app.mainloop()